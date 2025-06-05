import os
import pandas as pd
from openpyxl import load_workbook


class ExportaExcel:
    def __init__(self, gerenciador_rec, gerenciador_desp, resumos, caminho_base='./exportados'):
        self.gerenciador_rec = gerenciador_rec
        self.gerenciador_desp = gerenciador_desp
        self.resumos = resumos
        self.caminho_base = caminho_base
        os.makedirs(self.caminho_base, exist_ok=True)

    @staticmethod
    def converter_para_dataframe(objetos):
        return pd.DataFrame([{
            'Nome': obj.nome,
            'Tipo': obj.tipo,
            'Valor': obj.valor,
            'Categoria': obj.categoria,
            'Data': obj.data.strftime('%d/%m/%Y')
        } for obj in objetos])

    def exportar_receitas(self, receitas):
        if not receitas:
            print('Não há receitas para exportar!')
            return

        df_receitas = self.converter_para_dataframe(receitas)

        nome_arquivo = 'todas_receitas.xlsx'
        caminho_completo = os.path.join(self.caminho_base, nome_arquivo)
        try:
            df_receitas.to_excel(caminho_completo, index=False)
            print(f'Receitas exportadas para "{caminho_completo}" com sucesso!')
        except Exception as e:
            print(f'\033[31mErro ao exportar arquivo: {e}\033[m')

    def exportar_despesas(self, despesas):
        if not despesas:
            print('Não há despesas para exportar!')
            return

        df_despesas = self.converter_para_dataframe(despesas)

        nome_arquivo = 'todas_despesas.xlsx'
        caminho_completo = os.path.join(self.caminho_base, nome_arquivo)

        try:
            df_despesas.to_excel(caminho_completo, index=False)
            print(f'Despesas exportadas para "{caminho_completo}" com sucesso!')
        except Exception as e:
            print(f'\033[31mErro ao exportar arquivo: {e}\033[m')

    def exportar_receitas_periodo(self, ano, mes):
        receitas_mes = self.gerenciador_rec.filtrar_por_mes(ano, mes)

        if not receitas_mes:
            print(f'Não há receitas de {mes:02d}/{ano} para exportar!')
            return

        df_receitas_mes = self.converter_para_dataframe(receitas_mes)

        nome_arquivo = f'receitas_{ano}_{mes:02d}.xlsx'
        caminho_completo = os.path.join(self.caminho_base, nome_arquivo)

        try:
            df_receitas_mes.to_excel(caminho_completo, index=False)
            print(f'Receitas exportadas para "{caminho_completo}" com sucesso!')
        except Exception as e:
            print(f'\033[31mErro ao exportar arquivo: {e}\033[m')

    def exportar_despesas_periodo(self, ano, mes):
        despesas_mes = self.gerenciador_desp.filtrar_por_mes(ano, mes)

        if not despesas_mes:
            print(f'Não há despesas de {mes:02d}/{ano} para exportar!')
            return

        df_despesas_mes = self.converter_para_dataframe(despesas_mes)

        nome_arquivo = f'depesas_{ano}_{mes:02d}.xlsx'
        caminho_completo = os.path.join(self.caminho_base, nome_arquivo)

        try:
            df_despesas_mes.to_excel(caminho_completo, index=False)
            print(f'Despesas exportadas para "{caminho_completo}" com sucesso!')
        except Exception as e:
            print(f'\033[31mErro ao exportar arquivo: {e}\033[m')

    def exportar_resumo_mensal(self, ano, mes):
        if not self.resumos:
            print('Não há resumos ainda!')
            return
        elif (ano, mes) not in self.resumos.keys():
            print(f'Não há resumo de {mes:02d}/{ano} ainda!')
            return
        else:
            resumo = self.resumos[(ano, mes)]
            receitas_mes = self.gerenciador_rec.filtrar_por_mes(ano, mes)
            despesas_mes = self.gerenciador_desp.filtrar_por_mes(ano, mes)

            df_receitas_mes = self.converter_para_dataframe(receitas_mes)  # DataFrame de Receitas
            df_despesas_mes = self.converter_para_dataframe(despesas_mes)  # DataFrame de Despesas

            nome_arquivo = f'resumo_{ano}_{mes:02d}.xlsx'
            caminho_completo = os.path.join(self.caminho_base, nome_arquivo)

            try:
                # Criar o arquivo 'resumo_mensal.xlsx' com as planilhas de receita e despesa
                with pd.ExcelWriter(caminho_completo, engine='openpyxl') as writer:
                    df_receitas_mes.to_excel(writer, sheet_name='Receitas', index=False)
                    df_despesas_mes.to_excel(writer, sheet_name='Despesas', index=False)

                # Criar a planilha de Resumo Mensal com a biblioteca openpyxl
                wb = load_workbook(caminho_completo)
                planilha_resumo = wb.create_sheet('Resumo Mensal')

                planilha_resumo['A1'] = 'Resumo Mensal'
                planilha_resumo['A3'] = 'Valor Total Receitas'
                planilha_resumo['B3'] = resumo.total_receitas()
                planilha_resumo['A4'] = 'Valor Total Despesas'
                planilha_resumo['B4'] = resumo.total_despesas()
                planilha_resumo['A5'] = 'Saldo'
                planilha_resumo['B5'] = resumo.saldo_mensal()

                # Porcentagem por categoria
                planilha_resumo['A7'] = 'Porcentagens de Despesas por Categoria'
                planilha_resumo['A8'] = 'Categoria'
                planilha_resumo['B8'] = 'Porcentagem'
                linha = 9
                for categoria, porcentagem in resumo.porcentagens.items():
                    planilha_resumo[f'A{linha}'] = categoria
                    planilha_resumo[f'B{linha}'] = f'{porcentagem}%'
                    linha += 1

                # Análises
                planilha_resumo[f'A{linha + 1}'] = 'Categoria com menor gasto'
                planilha_resumo[f'B{linha + 1}'] = resumo.menor_categoria_despesa()[0]
                planilha_resumo[f'C{linha + 1}'] = resumo.menor_categoria_despesa()[1]

                planilha_resumo[f'A{linha + 2}'] = 'Categoria com maior gasto'
                planilha_resumo[f'B{linha + 2}'] = resumo.maior_categoria_despesa()[0]
                planilha_resumo[f'C{linha + 2}'] = resumo.maior_categoria_despesa()[1]

                planilha_resumo[f'A{linha + 4}'] = 'Menor Despesa'
                planilha_resumo[f'B{linha + 4}'] = resumo.menor_gasto()[0]
                planilha_resumo[f'C{linha + 4}'] = resumo.menor_gasto()[1]

                planilha_resumo[f'A{linha + 5}'] = 'Maior Despesa'
                planilha_resumo[f'B{linha + 5}'] = resumo.maior_gasto()[0]
                planilha_resumo[f'C{linha + 5}'] = resumo.maior_gasto()[1]

                # Salvando
                wb.save(caminho_completo)

                print(f'Resumo exportado para "{caminho_completo}" com sucesso!')
            except Exception as e:
                print(f'\033[31mErro ao exportar arquivo: {e}\033[m')